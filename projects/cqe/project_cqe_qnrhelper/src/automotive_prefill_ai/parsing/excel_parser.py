from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from automotive_prefill_ai.config.defaults import COLUMN_KEYWORDS, SHEET_HINT_KEYWORDS
from automotive_prefill_ai.models.questionnaire import (
    ParseResult,
    ParsedCell,
    ParsedQuestionRecord,
    ParsedSheet,
)


MAX_SCAN_ROWS = 80
MAX_SCAN_COLUMNS = 20
MAX_EMPTY_STREAK = 12
CATEGORY_HINTS = [
    "category",
    "section",
    "process element",
    "module",
    "類別",
    "章節",
    "項目",
]


@dataclass
class HeaderDetection:
    row_number: int | None
    columns: dict[str, int]
    score: int


class ExcelQuestionnaireParser:
    def parse(self, workbook_path: str | Path) -> ParseResult:
        workbook_path = Path(workbook_path)
        workbook = load_workbook(workbook_path, data_only=True)

        parsed_sheets: list[ParsedSheet] = []
        total_questions = 0

        for sheet in workbook.worksheets:
            parsed_sheet = self._parse_sheet(sheet)
            parsed_sheets.append(parsed_sheet)
            total_questions += parsed_sheet.question_count

        candidate_sheet_count = sum(1 for sheet in parsed_sheets if sheet.is_candidate_sheet)
        return ParseResult(
            workbook_name=workbook_path.name,
            total_sheets=len(parsed_sheets),
            candidate_sheet_count=candidate_sheet_count,
            question_count=total_questions,
            sheets=parsed_sheets,
        )

    def _parse_sheet(self, sheet: Worksheet) -> ParsedSheet:
        header = self._detect_header(sheet)
        is_candidate_sheet = header.score >= 2
        is_hidden = sheet.sheet_state != "visible"
        notes: list[str] = []
        questions: list[ParsedQuestionRecord] = []

        if is_hidden:
            notes.append("Sheet is hidden.")
        if not is_candidate_sheet:
            notes.append("No reliable questionnaire header detected.")
        elif header.row_number is None:
            notes.append("Candidate sheet detected but no header row was resolved.")
        else:
            questions = self._extract_questions(sheet, header)
            if not questions:
                notes.append("No question rows extracted after header detection.")

        return ParsedSheet(
            sheet_name=sheet.title,
            is_candidate_sheet=is_candidate_sheet,
            is_hidden=is_hidden,
            header_row=header.row_number,
            detected_columns=header.columns,
            row_count=sheet.max_row,
            question_count=sum(1 for item in questions if item.row_type == "question"),
            category_count=sum(1 for item in questions if item.row_type == "category"),
            notes=notes,
            questions=questions,
        )

    def _detect_header(self, sheet: Worksheet) -> HeaderDetection:
        best = HeaderDetection(row_number=None, columns={}, score=0)

        for row_number in range(1, min(sheet.max_row, MAX_SCAN_ROWS) + 1):
            row_values = [
                self._normalize_label(self._get_cell_value(sheet, row_number, column))
                for column in range(1, min(sheet.max_column, MAX_SCAN_COLUMNS) + 1)
            ]

            detected_columns: dict[str, int] = {}
            row_score = 0

            for semantic_name, keywords in COLUMN_KEYWORDS.items():
                match_index = self._find_keyword_index(row_values, keywords)
                if match_index is not None:
                    detected_columns[semantic_name] = match_index + 1
                    row_score += 1

            hint_bonus = sum(1 for value in row_values if any(hint in value for hint in SHEET_HINT_KEYWORDS))
            row_score += 1 if hint_bonus >= 2 else 0

            if row_score > best.score:
                best = HeaderDetection(row_number=row_number, columns=detected_columns, score=row_score)

        return best

    def _extract_questions(self, sheet: Worksheet, header: HeaderDetection) -> list[ParsedQuestionRecord]:
        question_column = header.columns.get("question")
        if question_column is None or header.row_number is None:
            return []

        records: list[ParsedQuestionRecord] = []
        empty_streak = 0

        for row_number in range(header.row_number + 1, sheet.max_row + 1):
            question_value = self._cell_text(self._get_cell_value(sheet, row_number, question_column))

            if not question_value:
                empty_streak += 1
                if empty_streak >= MAX_EMPTY_STREAK:
                    break
                continue

            empty_streak = 0
            row_type = self._classify_row_type(sheet, row_number, header.columns, question_value)
            if row_type == "ignored":
                continue

            cells = self._collect_source_cells(sheet, row_number, header.columns)
            record = ParsedQuestionRecord(
                sheet_name=sheet.title,
                row_number=row_number,
                question_id=self._extract_optional_value(sheet, row_number, header.columns.get("question_id")),
                original_question=question_value,
                normalized_question=self._normalize_question(question_value),
                row_type="category" if row_type == "category" else "question",
                candidate_score=self._extract_optional_value(sheet, row_number, header.columns.get("score")),
                candidate_answer=self._extract_optional_value(sheet, row_number, header.columns.get("answer")),
                candidate_owner=self._extract_optional_value(sheet, row_number, header.columns.get("owner")),
                confidence=self._row_confidence(question_value, header, row_type),
                source_cells=cells,
            )
            records.append(record)

        return records

    def _collect_source_cells(
        self,
        sheet: Worksheet,
        row_number: int,
        columns: dict[str, int],
    ) -> list[ParsedCell]:
        cells: list[ParsedCell] = []

        for column in sorted(set(columns.values())):
            value = self._cell_text(self._get_cell_value(sheet, row_number, column))
            if not value:
                continue
            cell = sheet.cell(row=row_number, column=column)
            cells.append(
                ParsedCell(
                    row=row_number,
                    column=column,
                    coordinate=cell.coordinate,
                    value=value,
                )
            )

        return cells

    def _extract_optional_value(self, sheet: Worksheet, row_number: int, column: int | None) -> str | None:
        if column is None:
            return None
        return self._cell_text(self._get_cell_value(sheet, row_number, column)) or None

    def _row_confidence(self, question_value: str, header: HeaderDetection, row_type: str) -> str:
        if row_type == "category":
            return "medium"
        if header.score >= 4 and len(question_value) >= 15:
            return "high"
        if header.score >= 2:
            return "medium"
        return "low"

    def _find_keyword_index(self, values: list[str], keywords: list[str]) -> int | None:
        for index, value in enumerate(values):
            if not value:
                continue
            for keyword in keywords:
                if keyword in value:
                    return index
        return None

    def _normalize_label(self, value: object) -> str:
        text = self._cell_text(value).lower()
        return re.sub(r"\s+", " ", text).strip()

    def _normalize_question(self, text: str) -> str:
        normalized = re.sub(r"\s+", " ", text).strip()
        normalized = re.sub(r"^[\d\.\-\)\(a-zA-Z一二三四五六七八九十]+[\s:：、\.]+", "", normalized)
        return normalized

    def _looks_like_question(self, text: str) -> bool:
        normalized = text.strip()
        if len(normalized) < 8:
            return False
        if normalized.lower() in {"yes", "no", "n/a"}:
            return False
        if normalized.count(" ") < 1 and len(normalized) < 12:
            return False
        return True

    def _classify_row_type(
        self,
        sheet: Worksheet,
        row_number: int,
        columns: dict[str, int],
        question_value: str,
    ) -> str:
        populated_values = [
            self._cell_text(self._get_cell_value(sheet, row_number, column))
            for column in sorted(set(columns.values()))
        ]
        non_empty_values = [value for value in populated_values if value]

        if self._looks_like_category(question_value, non_empty_values):
            return "category"
        if self._looks_like_question(question_value):
            return "question"
        return "ignored"

    def _looks_like_category(self, question_value: str, non_empty_values: list[str]) -> bool:
        normalized = question_value.strip().lower()
        if any(hint in normalized for hint in CATEGORY_HINTS):
            return True
        if len(non_empty_values) == 1 and len(normalized) <= 80 and "?" not in normalized and "？" not in normalized:
            return True
        if re.fullmatch(r"[a-zA-Z]?\d+(\.\d+)*\s*[-:：]?\s*.+", normalized) and len(non_empty_values) <= 2:
            return True
        return False

    def _get_cell_value(self, sheet: Worksheet, row_number: int, column: int) -> object:
        cell = sheet.cell(row=row_number, column=column)
        if not isinstance(cell, MergedCell):
            return cell.value

        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return sheet.cell(merged_range.min_row, merged_range.min_col).value
        return None

    def _cell_text(self, value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()
