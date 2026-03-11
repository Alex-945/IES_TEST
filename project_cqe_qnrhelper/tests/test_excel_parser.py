from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from automotive_prefill_ai.parsing.excel_parser import ExcelQuestionnaireParser


def test_excel_parser_extracts_question_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Assessment"
    sheet.append(["Item No", "Question", "Score", "Owner", "Answer"])
    sheet.append(["1", "Do you have an NTF handling process?", "", "", ""])
    sheet.append(["2", "Describe your APQP control method.", "", "", ""])
    workbook.save(workbook_path)

    result = ExcelQuestionnaireParser().parse(workbook_path)

    assert result.total_sheets == 1
    assert result.candidate_sheet_count == 1
    assert result.question_count == 2
    assert result.sheets[0].detected_columns["question"] == 2
    assert result.sheets[0].questions[0].normalized_question == "Do you have an NTF handling process?"


def test_excel_parser_detects_category_and_hidden_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample_hidden.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Main"
    sheet.append(["No", "Question", "Score", "Owner"])
    sheet.append(["P1", "Process Element 1: Project Management", "", ""])
    sheet.append(["1", "Do you define project responsibilities?", "", ""])

    hidden = workbook.create_sheet("HiddenData")
    hidden.sheet_state = "hidden"
    hidden.append(["Question", "Answer"])
    hidden.append(["Dummy hidden question content", ""])
    workbook.save(workbook_path)

    result = ExcelQuestionnaireParser().parse(workbook_path)

    assert result.total_sheets == 2
    assert result.candidate_sheet_count == 2
    assert result.sheets[0].category_count == 1
    assert result.sheets[0].question_count == 1
    assert result.sheets[0].questions[0].row_type == "category"
    assert result.sheets[1].is_hidden is True


def test_excel_parser_reads_merged_question_cells(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample_merged.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Merged"
    sheet.append(["Item", "Question", "Owner"])
    sheet.append(["1", "Do you keep documented evidence for APQP reviews?", "AQE"])
    sheet.merge_cells("B2:C2")
    workbook.save(workbook_path)

    result = ExcelQuestionnaireParser().parse(workbook_path)

    assert result.question_count == 1
    assert "documented evidence" in result.sheets[0].questions[0].original_question
